"""
Interfaz web para la simulacion de estacion de servicio.
Ejecutar: python app.py
Abrir: http://localhost:5000
"""

import io
from flask import Flask, request, render_template_string, send_file
from estacion_servicio import Config, Simulacion, _fmt
import html as html_mod

app = Flask(__name__)

SUB_COLS = ["Estado", "Tipo", "Llegada", "Ini espera", "Fin atencion"]
SUB_SUFFIXES = ["_estado", "_tipo", "_llegada", "_espera", "_fin"]


def _grupos_columnas(sim: Simulacion):
    nombres_surt = [s.nombre for s in sim.surtidores]
    nombres_gom = [s.nombre for s in sim.gomerias]
    nombres_acc = [s.nombre for s in sim.accesorios_list]
    return [
        ("evt", "Evento", ["N", "Evento", "Reloj"]),
        ("lleg", "Llegada cliente",
            ["RND1 lleg.", "RND2 lleg.", "T entre lleg. (s)", "Prox. llegada"]),
        ("ruta", "Tipo de servicio",
            ["RND servicio", "Tipo servicio"]),
        ("srv-carga", "Fin carga combustible",
            ["RND carga", "T carga (s)", "Fin carga"]),
        ("srv-acc", "Fin atencion accesorios",
            ["RND acces.", "T acces. (s)", "Fin acces."]),
        ("srv-gom", "Fin atencion gomeria",
            ["RND gomeria", "T gomeria (s)", "Fin gomeria"]),
        ("ruta2", "Evento extra",
            ["Cargo comb?", "RND post", "Que hace luego"]),
        ("servs", "Surtidores", nombres_surt + ["Cola surt."]),
        ("servs2", "Empleados gomeria", nombres_gom + ["Cola gom."]),
        ("servs3", "Accesorios", nombres_acc + ["Cola acces."]),
        ("stats", "Variables estadisticas",
            ["Max cola surt.", "Max cola gom.", "Max cola acces.",
             "Max T sist. (s)"]),
    ]


def _cli_ids_orden(ventana):
    cli_ids = set()
    for f in ventana:
        for k in f.keys():
            if k.endswith("_estado"):
                prefix = k[:-7]
                if prefix.startswith("C") and prefix[1:].isdigit():
                    cli_ids.add(int(prefix[1:]))
    return sorted(cli_ids)


def _ventana(sim: Simulacion):
    # La ventana ya fue aplicada por la simulacion: sim.filas contiene solo
    # las filas registradas (rango pedido + fila de estado final).
    return sim._rec_ini, sim.filas

CAMPOS = [
    ("n_clientes",   "Cantidad de clientes",               500,   "int",   "Simulacion",        "Cuantos clientes se generan"),
    ("n_iteraciones","Iteraciones (vacio = hasta el final)",  "",   "oint",  "Simulacion",        "Cortar la simulacion tras N iteraciones (eventos)"),
    ("semilla",      "Semilla (vacio = aleatoria)",          "",   "oint",  "Simulacion",        "Semilla del generador aleatorio"),
    ("desde",        "Mostrar desde fila",                    0,   "int",   "Visualizacion",     "Indice de la primera fila a mostrar"),
    ("cantidad",     "Cant. filas a mostrar (vacio = todas)","",   "oint",  "Visualizacion",     "Cantidad de filas a mostrar"),
    ("n_surtidores", "Surtidores",                            3,   "int",   "Servidores",        "Cantidad de surtidores de combustible"),
    ("n_gomerias",   "Empleados de gomeria",                  2,   "int",   "Servidores",        "Cantidad de empleados de gomeria"),
    ("n_accesorios", "Puestos de accesorios",                 1,   "int",   "Servidores",        "Cantidad de puestos de venta de accesorios"),
    ("lleg_media",   "Media (seg)",                         24.0,  "float", "Llegadas - Normal",         "Media de la distribucion normal (segundos)"),
    ("lleg_desv",    "Desv. estandar (seg)",                23.0,  "float", "Llegadas - Normal",         "Desviacion estandar (segundos)"),
    ("carga_a",      "A - minimo (seg)",                    45.0,  "float", "Carga combustible - Uniforme(A, B)", "Limite inferior en segundos"),
    ("carga_b",      "B - maximo (seg)",                    55.0,  "float", "Carga combustible - Uniforme(A, B)", "Limite superior en segundos"),
    ("gom_a",        "A - minimo (min)",                    10.0,  "float", "Gomeria - Uniforme(A, B)",  "Limite inferior en minutos"),
    ("gom_b",        "B - maximo (min)",                    26.0,  "float", "Gomeria - Uniforme(A, B)",  "Limite superior en minutos"),
    ("acc_a",        "A - minimo (min)",                     1.0,  "float", "Accesorios - Uniforme(A, B)", "Limite inferior en minutos"),
    ("acc_b",        "B - maximo (min)",                     5.0,  "float", "Accesorios - Uniforme(A, B)", "Limite superior en minutos"),
    ("p_carga",          "P(combustible)",                  0.80,  "float", "Probabilidades de ruteo", "Prob. de que un cliente vaya a cargar combustible"),
    ("p_no_carga_acc",   "P(no-carga -> accesorios)",       0.40,  "float", "Probabilidades de ruteo", "Si no carga: prob. de ir a accesorios (resto -> gomeria)"),
    ("p_post_carga_acc", "P(post-carga -> accesorios)",     0.30,  "float", "Probabilidades de ruteo", "Al terminar la carga: prob. de ir a accesorios"),
    ("p_post_carga_gom", "P(post-carga -> gomeria)",        0.20,  "float", "Probabilidades de ruteo", "Al terminar la carga: prob. de ir a gomeria (resto -> sale)"),
]


def _val(form, field_id, tipo, default):
    raw = form.get(field_id, "").strip()
    if raw == "":
        return default
    if tipo == "int":
        return int(raw)
    if tipo == "oint":
        return int(raw) if raw else None
    if tipo == "float":
        return float(raw)
    return raw


TEMPLATE = r"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Simulacion - Estacion de Servicio</title>
<style>
:root { color-scheme: light dark; --bg: #f5f6fa; --card: #fff; --border: #ddd;
        --text: #111; --text2: #555; --accent: #1976d2; --accent2: #ef6c00; }
@media (prefers-color-scheme: dark) {
    :root { --bg: #0e0e0e; --card: #1a1a1a; --border: #333;
            --text: #eee; --text2: #aaa; --accent: #64b5f6; --accent2: #ffb74d; }
}
*, *::before, *::after { box-sizing: border-box; }
body { font-family: 'Segoe UI', system-ui, sans-serif; margin: 0;
       background: var(--bg); color: var(--text); }
header { background: #263238; color: #fff; padding: 16px 24px; }
header h1 { margin: 0; font-size: 20px; font-weight: 600; }
header p  { margin: 4px 0 0; font-size: 13px; opacity: .7; }

.layout { display: flex; min-height: calc(100vh - 70px); }

.panel { width: 380px; min-width: 300px; background: var(--card);
         border-right: 1px solid var(--border); padding: 16px;
         overflow-y: auto; max-height: calc(100vh - 70px); }
.grupo-titulo {
    font-size: 12px; font-weight: 700; text-transform: uppercase;
    letter-spacing: .5px; color: var(--accent); margin: 18px 0 6px;
    border-bottom: 1px solid var(--border); padding-bottom: 4px;
}
.grupo-titulo:first-of-type { margin-top: 0; }
.campo { display: flex; align-items: center; gap: 8px; margin: 6px 0; }
.campo label { flex: 1; font-size: 13px; color: var(--text2); }
.campo input { width: 100px; padding: 5px 8px; border: 1px solid var(--border);
               border-radius: 4px; font-size: 13px; background: var(--bg);
               color: var(--text); text-align: right; }
.campo input:focus { outline: 2px solid var(--accent); border-color: transparent; }
.btn-row { margin-top: 16px; display: flex; gap: 8px; }
button {
    flex: 1; padding: 10px; border: none; border-radius: 6px; cursor: pointer;
    font-size: 14px; font-weight: 600; transition: .15s;
}
button[type=submit] { background: var(--accent); color: #fff; }
button[type=submit]:hover { filter: brightness(1.15); }
button[type=reset] { background: var(--border); color: var(--text); }
button.btn-excel { background: #1d6f42; color: #fff; }
button.btn-excel:hover { filter: brightness(1.15); }

.resultados { flex: 1; padding: 16px; display: flex; flex-direction: column;
             max-height: calc(100vh - 70px); overflow: hidden; }
.cards { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 16px; flex-shrink: 0; }
.card {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 8px; padding: 12px 16px; min-width: 200px;
    box-shadow: 0 1px 3px rgba(0,0,0,.06);
}
.card h3 { margin: 0 0 6px; font-size: 13px; color: var(--accent); text-transform: uppercase; letter-spacing: .3px; }
.card div { font-size: 13px; margin: 2px 0; }

.tabla-wrap {
    flex: 1; overflow: auto;
    border: 1px solid var(--border); border-radius: 6px;
    cursor: grab; user-select: none;
    min-height: 0;
}
.tabla-wrap.grabbing { cursor: grabbing; }
.tabla-wrap::-webkit-scrollbar { width: 8px; height: 8px; }
.tabla-wrap::-webkit-scrollbar-track { background: var(--bg); }
.tabla-wrap::-webkit-scrollbar-thumb { background: #888; border-radius: 4px; }

table { border-collapse: separate; border-spacing: 0; font-size: 12px; white-space: nowrap; width: max-content; }
th, td { padding: 4px 8px; text-align: right; border-right: 1px solid var(--border);
         border-bottom: 1px solid var(--border); }
thead th { position: sticky; z-index: 3; font-weight: 600; }
thead tr:first-child th { top: 0; }
thead tr:nth-child(2) th { top: 26px; }
tbody tr:nth-child(even) { background: rgba(128,128,128,.06); }
tbody tr:hover { background: rgba(25,118,210,.08); }

/* ultima fila de la simulacion: siempre visible al pie de la tabla */
tbody tr.sticky-bottom td {
    position: sticky; bottom: 0; z-index: 2;
    background: var(--card);
    border-top: 2px solid var(--accent);
    box-shadow: 0 -2px 4px rgba(0,0,0,.12);
    font-weight: 600;
}
tbody tr.sticky-bottom:hover td { background: var(--card); }

th.evt       { background: #37474f; color: #fff; }
th.lleg      { background: #1976d2; color: #fff; }
th.ruta      { background: #6a1b9a; color: #fff; }
th.srv-carga { background: #ef6c00; color: #fff; }
th.srv-acc   { background: #c2185b; color: #fff; }
th.srv-gom   { background: #2e7d32; color: #fff; }
th.ruta2     { background: #4a148c; color: #fff; }
th.servs     { background: #455a64; color: #fff; }
th.servs2    { background: #00695c; color: #fff; }
th.servs3    { background: #ad1457; color: #fff; }
th.stats     { background: #b71c1c; color: #fff; }
th.cli       { background: #5d4037; color: #fff; }

td.evento { text-align: left; font-weight: 600; }
td.ruta, td.ruta2 { text-align: left; }
td.cli    { text-align: left; font-family: ui-monospace, Consolas, monospace; font-size: 11px; }

.empty { text-align: center; padding: 80px 20px; color: var(--text2); flex: 1;
         display: flex; flex-direction: column; justify-content: center; align-items: center; }
.empty h2 { margin: 0 0 8px; font-size: 18px; }
</style>
</head><body>

<header>
    <h1>Simulacion - Estacion de Servicio</h1>
    <p>UTN FRC - Materia Simulacion - Todos los parametros son configurables</p>
</header>

<div class="layout">
<form class="panel" method="POST" action="/">
{% set ns = namespace(current_group=None) %}
{% for id, label, default, tipo, grupo, tooltip in campos %}
    {% if grupo != ns.current_group %}
        {% set ns.current_group = grupo %}
        <div class="grupo-titulo">{{ grupo }}</div>
    {% endif %}
    <div class="campo" title="{{ tooltip }}">
        <label for="{{ id }}">{{ label }}</label>
        <input id="{{ id }}" name="{{ id }}"
               type="{% if tipo in ('int','oint') %}number{% else %}text{% endif %}"
               {% if tipo in ('int','oint') %}step="1"{% endif %}
               {% if tipo == 'float' %}inputmode="decimal"{% endif %}
               value="{{ valores.get(id, default) }}"
               {% if tipo == 'oint' %}placeholder="---"{% endif %}
        >
    </div>
{% endfor %}

    <div class="btn-row">
        <button type="submit">Simular</button>
        <button type="reset">Reset</button>
    </div>
    <div class="btn-row">
        <button type="submit" formaction="/exportar" formnovalidate
                class="btn-excel">Exportar a Excel</button>
    </div>
</form>

<div class="resultados">
{% if not resultado %}
    <div class="empty">
        <h2>Configura los parametros y hace clic en Simular</h2>
        <p>Los valores por defecto corresponden al enunciado original del TP.</p>
    </div>
{% else %}
    <div class="cards">
        <div class="card">
            <h3>Simulacion</h3>
            <div><b>Clientes arribados:</b> {{ resultado.arribados }}</div>
            <div><b>Eventos totales:</b> {{ resultado.total_eventos }}</div>
            <div><b>Filas mostradas:</b> {{ resultado.filas_mostradas }}
                 (desde {{ resultado.desde }})</div>
            <div><b>Semilla:</b> {{ resultado.semilla }}</div>
            {% if resultado.truncado %}
            <div style="color: var(--accent2); font-size: 12px; margin-top: 4px;">
                Se registro solo una ventana de la corrida (+ estado final).
                Usa "Mostrar desde" / "Cant. filas" para navegar.
            </div>
            {% endif %}
        </div>
        <div class="card">
            <h3>Colas maximas</h3>
            <div><b>Surtidores:</b> {{ resultado.cola_max_surt }}</div>
            <div><b>Gomeria:</b> {{ resultado.cola_max_gom }}</div>
            <div><b>Accesorios:</b> {{ resultado.cola_max_acc }}</div>
        </div>
        <div class="card">
            <h3>Tiempo maximo en sistema</h3>
            <div><b>{{ resultado.t_max_s }}</b> s = <b>{{ resultado.t_max_m }}</b> min</div>
            <div>Cliente <b>C{{ resultado.t_max_cliente }}</b></div>
        </div>
        <div class="card">
            <h3>Parametros usados</h3>
            <div><b>Llegadas:</b> Normal(media={{ resultado.lleg_media }}s, desv={{ resultado.lleg_desv }}s)</div>
            <div><b>Carga:</b> U({{ resultado.carga_a }}, {{ resultado.carga_b }}) seg</div>
            <div><b>Gomeria:</b> U({{ resultado.gom_a }}, {{ resultado.gom_b }}) min</div>
            <div><b>Accesorios:</b> U({{ resultado.acc_a }}, {{ resultado.acc_b }}) min</div>
            <div><b>Servidores:</b> {{ resultado.n_surt }} surt - {{ resultado.n_gom }} gom - {{ resultado.n_acc }} acc</div>
        </div>
    </div>

    <div class="tabla-wrap">
    <table>
        <thead>
            <tr>{{ resultado.thead1 | safe }}</tr>
            <tr>{{ resultado.thead2 | safe }}</tr>
        </thead>
        <tbody>
            {{ resultado.tbody | safe }}
        </tbody>
    </table>
    </div>
{% endif %}
</div>
</div>

<script>
document.querySelectorAll('.tabla-wrap').forEach(function(el) {
    var startX, scrollLeft, down = false;
    el.addEventListener('mousedown', function(e) {
        down = true;
        el.classList.add('grabbing');
        startX = e.pageX - el.offsetLeft;
        scrollLeft = el.scrollLeft;
    });
    el.addEventListener('mouseleave', function() { down = false; el.classList.remove('grabbing'); });
    el.addEventListener('mouseup', function() { down = false; el.classList.remove('grabbing'); });
    el.addEventListener('mousemove', function(e) {
        if (!down) return;
        e.preventDefault();
        el.scrollLeft = scrollLeft - (e.pageX - el.offsetLeft - startX);
    });
});
</script>
</body></html>"""


def build_tabla(sim: Simulacion):
    cfg = sim.cfg

    inicio, ventana = _ventana(sim)

    grupos = _grupos_columnas(sim)
    cli_ids_orden = _cli_ids_orden(ventana)

    # thead row 1 (grupos)
    thead1 = ""
    for clave, titulo, cols in grupos:
        thead1 += (f'<th class="{clave}" colspan="{len(cols)}">'
                   f'{html_mod.escape(titulo)}</th>')
    for cid in cli_ids_orden:
        thead1 += (f'<th class="cli" colspan="{len(SUB_COLS)}">'
                   f'Cliente {cid}</th>')

    # thead row 2 (columnas individuales)
    thead2 = ""
    for clave, _, cols in grupos:
        for c in cols:
            thead2 += f'<th class="{clave}">{html_mod.escape(c)}</th>'
    for cid in cli_ids_orden:
        for sc in SUB_COLS:
            thead2 += f'<th class="cli">{html_mod.escape(sc)}</th>'

    # tbody
    filas_html = []
    ultima_idx = len(ventana) - 1
    for i, f in enumerate(ventana):
        celdas = []
        for clave, _, cols in grupos:
            for c in cols:
                extra = " evento" if c == "Evento" else ""
                celdas.append(
                    f'<td class="{clave}{extra}">{_fmt(f.get(c))}</td>')
        for cid in cli_ids_orden:
            prefix = f"C{cid}"
            for suffix in SUB_SUFFIXES:
                key = f"{prefix}{suffix}"
                celdas.append(f'<td class="cli">{_fmt(f.get(key, ""))}</td>')
        tr_class = ' class="sticky-bottom"' if i == ultima_idx else ""
        filas_html.append(f"<tr{tr_class}>" + "".join(celdas) + "</tr>")

    tmax_min = sim.stats.tiempo_max_sistema / 60

    return {
        "thead1": thead1,
        "thead2": thead2,
        "tbody": "\n".join(filas_html),
        "arribados": sim.arribados,
        "total_eventos": sim.n_evento,
        "filas_mostradas": len(ventana),
        "desde": inicio,
        "hasta": inicio + len(ventana) - 1,
        "truncado": sim.truncado,
        "semilla": cfg.semilla if cfg.semilla is not None else "aleatoria",
        "cola_max_surt": sim.stats.cola_max_surtidor,
        "cola_max_gom": sim.stats.cola_max_gomeria,
        "cola_max_acc": sim.stats.cola_max_accesorios,
        "t_max_s": f"{sim.stats.tiempo_max_sistema:.2f}",
        "t_max_m": f"{tmax_min:.2f}",
        "t_max_cliente": sim.stats.id_cliente_max_tiempo,
        "lleg_media": cfg.lleg_media,
        "lleg_desv": cfg.lleg_desv,
        "carga_a": cfg.carga_a,
        "carga_b": cfg.carga_b,
        "gom_a": cfg.gom_a,
        "gom_b": cfg.gom_b,
        "acc_a": cfg.acc_a,
        "acc_b": cfg.acc_b,
        "n_surt": cfg.n_surtidores,
        "n_gom": cfg.n_gomerias,
        "n_acc": cfg.n_accesorios,
    }


def _cfg_desde_form(f) -> Config:
    return Config(
        n_clientes=_val(f, "n_clientes", "int", 500),
        n_iteraciones=_val(f, "n_iteraciones", "oint", None),
        desde=_val(f, "desde", "int", 0),
        cantidad=_val(f, "cantidad", "oint", None),
        semilla=_val(f, "semilla", "oint", None),
        n_surtidores=_val(f, "n_surtidores", "int", 3),
        n_gomerias=_val(f, "n_gomerias", "int", 2),
        n_accesorios=_val(f, "n_accesorios", "int", 1),
        lleg_media=_val(f, "lleg_media", "float", 24.0),
        lleg_desv=_val(f, "lleg_desv", "float", 23.0),
        carga_a=_val(f, "carga_a", "float", 45.0),
        carga_b=_val(f, "carga_b", "float", 55.0),
        gom_a=_val(f, "gom_a", "float", 10.0),
        gom_b=_val(f, "gom_b", "float", 26.0),
        acc_a=_val(f, "acc_a", "float", 1.0),
        acc_b=_val(f, "acc_b", "float", 5.0),
        p_carga=_val(f, "p_carga", "float", 0.80),
        p_no_carga_acc=_val(f, "p_no_carga_acc", "float", 0.40),
        p_post_carga_acc=_val(f, "p_post_carga_acc", "float", 0.30),
        p_post_carga_gom=_val(f, "p_post_carga_gom", "float", 0.20),
    )


def _num(v: float) -> str:
    """Formatea un numero sin '.0' innecesario (12.0 -> '12', 0.8 -> '0.8')."""
    return f"{v:g}"


def _filas_formulas(cfg: Config):
    """Fuente unica de verdad de las formulas del modelo.
    Devuelve filas: (categoria, calculo, formula general, con parametros, detalle).
    """
    p_c = cfg.p_carga
    p_acc = p_c + (1 - p_c) * cfg.p_no_carga_acc
    p_post_acc = cfg.p_post_carga_acc
    p_post_gom = p_post_acc + cfg.p_post_carga_gom
    return [
        ("Llegadas",
         "Tiempo entre llegadas ~ Normal(media, desv)",
         "z = raiz(-2*ln(RND1)) * cos(2*pi*RND2) ;  T = media + desv*z",
         f"T = {_num(cfg.lleg_media)} + {_num(cfg.lleg_desv)}*z   (minimo 0.1 s)",
         "Metodo de Box-Muller. RND1, RND2 ~ U(0,1). Se toma max(0.1, T)."),
        ("Llegadas",
         "Proxima llegada",
         "Prox_llegada = Reloj + T",
         "-",
         "Se agenda el proximo arribo al generar el actual."),
        ("Carga combustible",
         "Duracion de carga ~ Uniforme(a, b)",
         "T = a + (b - a) * RND",
         f"T = {_num(cfg.carga_a)} + {_num(cfg.carga_b - cfg.carga_a)}*RND   [s]",
         f"a = {_num(cfg.carga_a)} s, b = {_num(cfg.carga_b)} s. RND ~ U(0,1)."),
        ("Gomeria",
         "Duracion de gomeria ~ Uniforme(a, b)  [min -> s]",
         "T = (a + (b - a) * RND) * 60",
         f"T = ({_num(cfg.gom_a)} + {_num(cfg.gom_b - cfg.gom_a)}*RND) * 60   [s]",
         f"a = {_num(cfg.gom_a)} min, b = {_num(cfg.gom_b)} min. Se pasa a segundos."),
        ("Accesorios",
         "Duracion de accesorios ~ Uniforme(a, b)  [min -> s]",
         "T = (a + (b - a) * RND) * 60",
         f"T = ({_num(cfg.acc_a)} + {_num(cfg.acc_b - cfg.acc_a)}*RND) * 60   [s]",
         f"a = {_num(cfg.acc_a)} min, b = {_num(cfg.acc_b)} min. Se pasa a segundos."),
        ("Ruteo - Tipo de servicio",
         "Seleccion con 1 RND y probabilidades acumuladas",
         "Carga si RND < p_c ; Accesorios si RND < p_c+(1-p_c)*p_nc ; Gomeria si no",
         f"Carga si RND < {_num(p_c)} ; Accesorios si RND < {_num(p_acc)} ; "
         f"Gomeria si RND >= {_num(p_acc)}",
         f"p_c = {_num(p_c)} (carga), p_nc = {_num(cfg.p_no_carga_acc)} "
         f"(accesorios entre los que no cargan)."),
        ("Ruteo - Post carga",
         "Que hace el cliente al terminar de cargar",
         "Accesorios si RND < p_a ; Gomeria si RND < p_a+p_g ; Se retira si no",
         f"Accesorios si RND < {_num(p_post_acc)} ; Gomeria si RND < {_num(p_post_gom)} ; "
         f"Se retira si RND >= {_num(p_post_gom)}",
         f"p_a = {_num(cfg.p_post_carga_acc)}, p_g = {_num(cfg.p_post_carga_gom)}."),
        ("Atencion",
         "Fin de atencion de un servidor",
         "Fin = Reloj + T",
         "-",
         "T es la duracion sorteada del servicio correspondiente."),
        ("Estadisticas",
         "Tiempo del cliente en el sistema",
         "T_sist = Salida - Llegada",
         "-",
         "Se guarda el maximo y el cliente que lo produjo."),
        ("Estadisticas",
         "Cola maxima por sector",
         "max_cola = max(max_cola, cantidad_en_cola)",
         "-",
         "Se actualiza tras cada evento para surtidores, gomeria y accesorios."),
    ]


def _hoja_formulas(wb, cfg: Config):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Formulas")
    encabezados = ["Categoria", "Calculo", "Formula general",
                   "Con parametros actuales", "Detalle"]

    # titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    t = ws.cell(row=1, column=1, value="Formulas utilizadas en el modelo")
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="37474F")
    t.alignment = Alignment(horizontal="center", vertical="center")

    # encabezados
    hfill = PatternFill("solid", fgColor="455A64")
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # filas
    wrap = Alignment(vertical="top", wrap_text=True)
    mono = Font(name="Consolas")
    for i, fila in enumerate(_filas_formulas(cfg), start=3):
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = wrap
            if j in (3, 4):  # columnas de formula en monoespaciada
                c.font = mono

    anchos = [22, 40, 52, 46, 50]
    from openpyxl.utils import get_column_letter
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    return ws


def build_xlsx(sim: Simulacion) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    inicio, ventana = _ventana(sim)
    grupos = _grupos_columnas(sim)
    cli_ids_orden = _cli_ids_orden(ventana)

    # colores por grupo (mismos que la web, sin '#')
    colores = {
        "evt": "37474F", "lleg": "1976D2", "ruta": "6A1B9A",
        "srv-carga": "EF6C00", "srv-acc": "C2185B", "srv-gom": "2E7D32",
        "ruta2": "4A148C", "servs": "455A64", "servs2": "00695C",
        "servs3": "AD1457", "stats": "B71C1C", "cli": "5D4037",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Simulacion"

    white = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # fila 1: grupos (celdas combinadas)  |  fila 2: columnas
    col = 1
    for clave, titulo, cols in grupos:
        fill = PatternFill("solid", fgColor=colores[clave])
        ini = col
        for c in cols:
            cell = ws.cell(row=2, column=col, value=c)
            cell.fill = fill
            cell.font = white
            cell.alignment = center
            col += 1
        g = ws.cell(row=1, column=ini, value=titulo)
        g.fill = fill
        g.font = white
        g.alignment = center
        if col - 1 > ini:
            ws.merge_cells(start_row=1, start_column=ini,
                           end_row=1, end_column=col - 1)

    fill_cli = PatternFill("solid", fgColor=colores["cli"])
    for cid in cli_ids_orden:
        ini = col
        for sc in SUB_COLS:
            cell = ws.cell(row=2, column=col, value=sc)
            cell.fill = fill_cli
            cell.font = white
            cell.alignment = center
            col += 1
        g = ws.cell(row=1, column=ini, value=f"Cliente {cid}")
        g.fill = fill_cli
        g.font = white
        g.alignment = center
        ws.merge_cells(start_row=1, start_column=ini,
                       end_row=1, end_column=col - 1)

    # datos
    r = 3
    for f in ventana:
        col = 1
        for _, _, cols in grupos:
            for c in cols:
                ws.cell(row=r, column=col, value=_celda_valor(f.get(c)))
                col += 1
        for cid in cli_ids_orden:
            prefix = f"C{cid}"
            for suffix in SUB_SUFFIXES:
                ws.cell(row=r, column=col,
                        value=_celda_valor(f.get(f"{prefix}{suffix}", "")))
                col += 1
        r += 1

    # fila final resaltada
    ultima = r - 1
    hi = PatternFill("solid", fgColor="FFF3CD")
    for c in range(1, col):
        cell = ws.cell(row=ultima, column=c)
        cell.fill = hi
        cell.font = Font(bold=True)

    ws.freeze_panes = "A3"
    for c in range(1, col):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # hoja adicional con las formulas del modelo
    _hoja_formulas(wb, sim.cfg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _celda_valor(v):
    """Devuelve numeros como numeros y el resto como texto para Excel."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v)


@app.route("/exportar", methods=["POST"])
def exportar():
    f = request.form
    cfg = _cfg_desde_form(f)

    sim = Simulacion(cfg)
    sim.ejecutar()
    buf = build_xlsx(sim)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="simulacion_estacion.xlsx",
    )


@app.route("/", methods=["GET", "POST"])
def index():
    resultado = None
    valores = {}

    if request.method == "POST":
        f = request.form
        for campo_id, _, default, tipo, _, _ in CAMPOS:
            raw = f.get(campo_id, "").strip()
            valores[campo_id] = raw if raw else default

        cfg = _cfg_desde_form(f)

        sim = Simulacion(cfg)
        sim.ejecutar()
        resultado = build_tabla(sim)

    return render_template_string(
        TEMPLATE,
        campos=CAMPOS,
        valores=valores,
        resultado=resultado,
    )


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    print(f"  Abri http://localhost:{port} en tu navegador")
    app.run(debug=True, port=port)
